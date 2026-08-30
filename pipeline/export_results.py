import csv
import json
from glob import glob
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
JSON_DIR = ROOT / "scenario_json_details"
OUT_CSV = ROOT / "climate_ipcc_pipeline_result_optimized.csv"
OUT_XLSX = ROOT / "climate_ipcc_pipeline_result_optimized.xlsx"


def load_items():
    items = []
    for path in sorted(glob(str(JSON_DIR / "*.json"))):
        try:
            with open(path, "r", encoding="utf-8") as f:
                items.append(json.load(f))
        except Exception:
            continue
    return items


def clean_text(value):
    if value is None:
        return "N/A"
    text = str(value).strip()
    return text if text else "N/A"


def format_timeseries(timeseries):
    if not isinstance(timeseries, list) or not timeseries:
        return "N/A"
    parts = []
    for item in timeseries:
        if isinstance(item, dict):
            year = clean_text(item.get("year"))
            variable = clean_text(item.get("variable"))
            value = clean_text(item.get("value"))
            context = clean_text(item.get("context"))
            parts.append(f"[{year}] {variable}: {value} ({context})")
    return " || ".join(parts) if parts else "N/A"


def round1_cell(debate, model_name):
    if not debate:
        return "Not Checked"
    data = (debate.get("round1") or {}).get(model_name, {})
    if not isinstance(data, dict):
        return "Not Checked"
    if "error" in data:
        return f"Error: {data['error']}"
    parts = []
    for key, label in [
        ("scenario_valid", "S"),
        ("output_valid", "O"),
        ("summary_valid", "Sum"),
        ("ar6_mapping_valid", "AR6"),
        ("timeseries_valid", "TS"),
    ]:
        val = data.get(key)
        if isinstance(val, bool):
            parts.append(f"{label}={'Y' if val else 'N'}")
    reasoning = clean_text(data.get("scenario_reasoning", ""))
    if reasoning != "N/A":
        parts.append(reasoning)
    return " | ".join(parts) if parts else "Not Checked"


def round2_cell(debate, model_name):
    if not debate:
        return "Not Checked"
    data = (debate.get("round2") or {}).get(model_name, {})
    if not isinstance(data, dict):
        return "Not Checked"
    if "error" in data:
        return f"Error: {data['error']}"
    uj = data.get("updated_judgment", data)
    parts = []
    for key, label in [
        ("scenario_valid", "S"),
        ("output_valid", "O"),
        ("summary_valid", "Sum"),
        ("ar6_mapping_valid", "AR6"),
        ("timeseries_valid", "TS"),
    ]:
        val = uj.get(key)
        if isinstance(val, bool):
            parts.append(f"{label}={'Y' if val else 'N'}")
    reasoning = clean_text(uj.get("final_reasoning", ""))
    if reasoning != "N/A":
        parts.append(reasoning)
    return " | ".join(parts) if parts else "Not Checked"


def debate_outcome(debate):
    if not debate:
        return "Not Checked"
    if debate.get("debate_skipped"):
        return f"Debate Skipped: {clean_text(debate.get('skip_reason'))}"
    if debate.get("scenario_valid") is not True:
        return "INVALID SCENARIO"
    if debate.get("extraction_complete") is False:
        return "SCENARIO VALID / EXTRACTION INCOMPLETE"
    issues = debate.get("issues_found", [])
    if not issues:
        return "VALID / NO ISSUES"
    if debate.get("auto_fix_applied"):
        return "VALID / AUTO-FIX APPLIED"
    return "VALID / ISSUES REMAIN"


def issue_summary(debate):
    if not debate:
        return "N/A"
    issues = debate.get("issues_found", [])
    if not issues:
        return "None"
    return " | ".join(
        f"[{clean_text(i.get('field'))}] {clean_text(i.get('problem'))}"
        for i in issues
        if isinstance(i, dict)
    ) or "None"


def screen_status(item):
    screen = item.get("scenario_screen", {})
    if screen.get("has_scenario") is False:
        return f"Filtered: {clean_text(screen.get('reason'))}"
    return "Passed"


def paper_row_placeholder(item):
    screen = item.get("scenario_screen", {})
    label = "No Valid Scenario"
    if screen.get("has_scenario") is False:
        label = f"No Scenario (filtered: {clean_text(screen.get('reason'))})"

    basic = item.get("basic_info", {})
    return [
        clean_text(item.get("file_name")),
        clean_text(item.get("expert")),
        clean_text(basic.get("title")),
        clean_text(basic.get("authors")),
        clean_text(basic.get("journal")),
        clean_text(basic.get("doi")),
        clean_text(basic.get("pub_year")),
        clean_text(item.get("location")),
        screen_status(item),
        label,
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
    ]


def build_rows(items):
    rows = []
    for item in items:
        basic = item.get("basic_info", {})
        scenarios = (item.get("scenarios_data") or {}).get("scenarios", [])
        if not scenarios:
            rows.append(paper_row_placeholder(item))
            continue

        for scenario in scenarios:
            debate = scenario.get("debate_validation") or {}
            rows.append([
                clean_text(item.get("file_name")),
                clean_text(item.get("expert")),
                clean_text(basic.get("title")),
                clean_text(basic.get("authors")),
                clean_text(basic.get("journal")),
                clean_text(basic.get("doi")),
                clean_text(basic.get("pub_year")),
                clean_text(item.get("location")),
                screen_status(item),
                clean_text(scenario.get("scenario_name")),
                clean_text(scenario.get("scenario_description")),
                clean_text(scenario.get("model_name")),
                clean_text(scenario.get("model_spatial_resolution")),
                clean_text(scenario.get("model_time_span")),
                clean_text(scenario.get("time_horizon")),
                clean_text(scenario.get("ar6_category")),
                clean_text(scenario.get("ar6_temperature_target")),
                clean_text(scenario.get("ar6_net_zero_year")),
                clean_text(scenario.get("ar6_carbon_budget")),
                clean_text(scenario.get("scenario_context_text")),
                clean_text(scenario.get("output_variable")),
                clean_text(scenario.get("output_value")),
                clean_text(scenario.get("output_description")),
                clean_text(scenario.get("output_context_text")),
                format_timeseries(scenario.get("key_variables_timeseries")),
                clean_text(scenario.get("structured_summary")),
                round1_cell(debate, "openai"),
                round1_cell(debate, "gemini"),
                round1_cell(debate, "kimi"),
                round2_cell(debate, "openai"),
                round2_cell(debate, "gemini"),
                round2_cell(debate, "kimi"),
                debate_outcome(debate),
                issue_summary(debate),
                "Yes" if debate.get("auto_fix_applied") else "No",
            ])
    return rows


HEADERS = [
    "File Name",
    "Expert Category",
    "Title",
    "Authors",
    "Journal",
    "DOI",
    "Year",
    "Study Location",
    "Screen Status",
    "Scenario Name",
    "Scenario Description",
    "Model Name",
    "Model Spatial Resolution",
    "Model Time Span",
    "Time Horizon",
    "AR6 Category",
    "AR6 Temperature Target / Peak Warming",
    "AR6 Net Zero Year",
    "AR6 Carbon Budget",
    "Scenario Context (Original Text)",
    "Output Variable",
    "Output Value",
    "Output Description",
    "Output Context (Original Text)",
    "Key Variables Timeseries",
    "Structured Summary",
    "OpenAI Round1 Review",
    "Gemini Round1 Review",
    "Kimi Round1 Review",
    "OpenAI Round2 Review",
    "Gemini Round2 Review",
    "Kimi Round2 Review",
    "Debate Outcome",
    "Issues Found",
    "Auto-Fix Applied",
]


def write_csv(rows):
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(rows)


def write_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    bad_fill = PatternFill("solid", fgColor="F4CCCC")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    good_fill = PatternFill("solid", fgColor="D9EAD3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_map = {
        "A": 42, "B": 18, "C": 40, "D": 26, "E": 24, "F": 18, "G": 10, "H": 24, "I": 28,
        "J": 24, "K": 48, "L": 22, "M": 18, "N": 18, "O": 14, "P": 14, "Q": 18, "R": 14,
        "S": 16, "T": 56, "U": 22, "V": 18, "W": 40, "X": 56, "Y": 28, "Z": 28, "AA": 28,
        "AB": 36, "AC": 36, "AD": 36, "AE": 24, "AF": 48, "AG": 14,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    debate_outcome_col = HEADERS.index("Debate Outcome") + 1
    scenario_name_col = HEADERS.index("Scenario Name") + 1

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

        outcome_cell = row[debate_outcome_col - 1]
        scenario_cell = row[scenario_name_col - 1]
        outcome = str(outcome_cell.value or "")
        scenario_name = str(scenario_cell.value or "")

        if scenario_name.startswith("No Scenario") or scenario_name == "No Valid Scenario":
            scenario_cell.fill = bad_fill
        if outcome.startswith("INVALID"):
            outcome_cell.fill = bad_fill
        elif "INCOMPLETE" in outcome:
            outcome_cell.fill = warn_fill
        elif outcome.startswith("VALID"):
            outcome_cell.fill = good_fill

    wb.save(OUT_XLSX)


def main():
    items = load_items()
    rows = build_rows(items)
    write_csv(rows)
    write_xlsx(rows)
    print(f"JSON files read: {len(items)}")
    print(f"Rows written: {len(rows)}")
    print(f"CSV: {OUT_CSV}")
    print(f"XLSX: {OUT_XLSX}")


if __name__ == "__main__":
    main()
