import os
import glob
import json
import csv
import textwrap
import re
import hashlib
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional, Tuple, Dict, Any, List

import pdfplumber
import pandas as pd
from openai import OpenAI
from tqdm import tqdm
from dotenv import load_dotenv


load_dotenv()

# ---------------------------------------------------------------------------
# Configuration. Nothing here is hard-coded to the authors' machine or accounts.
# Copy .env.example to .env and fill in your own values, or export the same names
# as environment variables. See README "Configuration".
# ---------------------------------------------------------------------------

def _require(name, hint):
    """Read a required setting, or fail immediately with an actionable message.

    Treats an untouched .env.example placeholder as missing, so a forgotten field
    fails here rather than as an authentication error several minutes in.
    """
    val = (os.getenv(name) or "").strip()
    if not val or val.lower().startswith(("insert your", "insert the")):
        raise SystemExit(
            f"\nMissing required setting: {name}\n"
            f"  {hint}\n"
            f"  Copy .env.example to .env and replace the placeholder, "
            f"or export {name}=... first.\n")
    return val


def _clean(name):
    """Read an optional setting, treating an untouched placeholder as unset."""
    v = (os.getenv(name) or "").strip()
    return None if v.lower().startswith(("insert your", "insert the")) else (v or None)


# Extraction and screening model
API_KEY = _clean("OPENAI_API_KEY")
BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
OPENAI_EXTRACT_MODEL = os.getenv("OPENAI_EXTRACT_MODEL", "gpt-4o")

# Debate panel member 2
GEMINI_API_KEY = _clean("GEMINI_API_KEY") or _clean("GOOGLE_API_KEY")
GEMINI_BASE_URL = os.getenv("GEMINI_BASE_URL", "https://generativelanguage.googleapis.com/v1beta/openai/")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-pro")

# Debate panel member 3
KIMI_API_KEY = _clean("KIMI_API_KEY")
KIMI_BASE_URL = os.getenv("KIMI_BASE_URL", "https://api.moonshot.cn/v1")
KIMI_MODEL = os.getenv("KIMI_MODEL", "kimi-k2-0711-preview")

# Debate panel member 1, defaults to the extraction account
GPT4O_API_KEY = _clean("GPT4O_API_KEY") or API_KEY
GPT4O_BASE_URL = os.getenv("GPT4O_BASE_URL", BASE_URL)
GPT4O_MODEL = os.getenv("GPT4O_MODEL", "gpt-4o")

# Local PDF corpus. Point this at your own library, e.g. a Zotero storage
# directory or the output of harvest_papers.py.
FOLDER_PATH = os.getenv("PDF_CORPUS_DIR") or _require(
    "PDF_CORPUS_DIR", "Directory holding the PDFs to process.")
OUTPUT_CSV = "climate_ipcc_pipeline_result.csv"
OUTPUT_DIR = "scenario_json_details"
SAVE_EVERY_N = 5

# Optional limit when running tests (set env SCENARIO_MAX_FILES=10, etc.)
MAX_FILES_OVERRIDE = int(os.getenv("SCENARIO_MAX_FILES", "0") or 0)

# Concurrency: papers processed in parallel (SCENARIO_WORKERS=10 by default)
MAX_WORKERS = int(os.getenv("SCENARIO_WORKERS", "20"))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AR6_DIR = os.path.join(BASE_DIR, "AR6_scenarios")
IPCC_REPORTS_DIR = os.path.join(BASE_DIR, "ipcc_reports")

HARDCODED_AR6_MAPPING = {
    # SSP scenarios (CMIP6)
    "ssp119": {"ar6_category": "C1", "ar6_temperature_target": "1.5°C", "ar6_net_zero_year": "2050", "ar6_carbon_budget": "500 Gt CO2"},
    "ssp126": {"ar6_category": "C3", "ar6_temperature_target": "1.8°C", "ar6_net_zero_year": "2070", "ar6_carbon_budget": "900 Gt CO2"},
    "ssp245": {"ar6_category": "C5", "ar6_temperature_target": "2.7°C", "ar6_net_zero_year": "N/A", "ar6_carbon_budget": "N/A"},
    "ssp370": {"ar6_category": "C7", "ar6_temperature_target": "3.6°C", "ar6_net_zero_year": "N/A", "ar6_carbon_budget": "N/A"},
    "ssp434": {"ar6_category": "C4", "ar6_temperature_target": "2.1°C", "ar6_net_zero_year": "2080", "ar6_carbon_budget": "N/A"},
    "ssp460": {"ar6_category": "C6", "ar6_temperature_target": "2.9°C", "ar6_net_zero_year": "N/A", "ar6_carbon_budget": "N/A"},
    "ssp534os": {"ar6_category": "C2", "ar6_temperature_target": "1.6°C", "ar6_net_zero_year": "2055", "ar6_carbon_budget": "700 Gt CO2"},
    "ssp585": {"ar6_category": "C8", "ar6_temperature_target": ">4°C (4.4°C)", "ar6_net_zero_year": "N/A", "ar6_carbon_budget": "N/A"},

    "rcp26": {"ar6_category": "C3 (RCP equivalent)", "ar6_temperature_target": "~1.8°C", "ar6_net_zero_year": "~2070", "ar6_carbon_budget": "~900 Gt CO2"},
    "rcp45": {"ar6_category": "C5 (RCP equivalent)", "ar6_temperature_target": "~2.7°C", "ar6_net_zero_year": "N/A", "ar6_carbon_budget": "N/A"},
    "rcp60": {"ar6_category": "C6 (RCP equivalent)", "ar6_temperature_target": "~3.1°C", "ar6_net_zero_year": "N/A", "ar6_carbon_budget": "N/A"},
    "rcp85": {"ar6_category": "C8 (RCP equivalent)", "ar6_temperature_target": ">4°C (4.3°C)", "ar6_net_zero_year": "N/A", "ar6_carbon_budget": "N/A"},
}


def normalize_key(value: Optional[str]) -> str:
    if not value:
        return ""
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def clean_value(value: Any) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, float) and pd.isna(value):
        return "N/A"
    text = str(value).strip()
    return text if text else "N/A"


def is_blank_like(value: Any) -> bool:
    """Treat non-string model outputs safely when checking whether a field is empty."""
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, (list, dict, tuple, set)):
        return len(value) == 0
    text = str(value).strip()
    return text == "" or text == "N/A"


def find_ar6_metadata_file() -> Optional[str]:
    if not os.path.isdir(AR6_DIR):
        return None
    for root, _, files in os.walk(AR6_DIR):
        for fname in files:
            if fname.endswith("metadata_indicators_v1.1.xlsx"):
                return os.path.join(root, fname)
    return None


def build_ar6_reference_summary(max_entries: int = 311) -> Tuple[str, Dict[str, Dict[str, Any]]]:
    metadata_path = find_ar6_metadata_file()
    if not metadata_path:
        return "Unable to locate AR6 metadata workbook. Only extract scenarios if they explicitly match IPCC-style definitions (SSP/RCP, mitigation policies, carbon budgets, etc.).", {}

    try:
        df = pd.read_excel(metadata_path, sheet_name="meta_Ch3vetted_withclimate")
    except Exception as exc:
        print(f"⚠️ Cannot read AR6 metadata: {exc}")
        return "AR6 reference unavailable. Stick to strict policy/socio-economic/forcing scenario definitions (SSP/RCP, Net Zero, carbon budgets).", {}

    selected_cols = [
        "Model",
        "Scenario",
        "Category",
        "Category_name",
        "Policy_category_name",
        "Ssp_family",
        "Technology_category_name",
        "Year of peak CO2 Emissions (Harm-Infilled)",
        "Year of netzero CO2 emissions (Harm-Infilled) Table SPM2",
        "Median peak warming (MAGICCv7.5.3)",
        "Cumulative net CO2 (2020-2100, Gt CO2) (Harm-Infilled)",
    ]

    missing_cols = [c for c in selected_cols if c not in df.columns]
    if missing_cols:
        print(f"⚠️ AR6 data is missing columns: {missing_cols}")
    available_cols = [c for c in selected_cols if c in df.columns]
    df_subset = df[available_cols].fillna("N/A")

    lookup: Dict[str, Dict[str, Any]] = {}
    entries: List[str] = []

    for _, row in df_subset.iterrows():
        scenario_name = str(row.get("Scenario", "")).strip()
        if not scenario_name:
            continue
        norm = normalize_key(scenario_name)
        if norm not in lookup:
            lookup[norm] = {
                "scenario": scenario_name,
                "category": clean_value(row.get("Category", "N/A")),
                "category_name": clean_value(row.get("Category_name", "N/A")),
                "policy": clean_value(row.get("Policy_category_name", "N/A")),
                "ssp": clean_value(row.get("Ssp_family", "N/A")),
                "technology": clean_value(row.get("Technology_category_name", "N/A")),
                "peak_year": clean_value(row.get("Year of peak CO2 Emissions (Harm-Infilled)", "N/A")),
                "net_zero_year": clean_value(row.get("Year of netzero CO2 emissions (Harm-Infilled) Table SPM2", "N/A")),
                "median_peak_warming": clean_value(row.get("Median peak warming (MAGICCv7.5.3)", "N/A")),
                "carbon_budget": clean_value(row.get("Cumulative net CO2 (2020-2100, Gt CO2) (Harm-Infilled)", "N/A")),
            }

        if len(entries) < max_entries:
            entry = (
                f"- {scenario_name} | Category: {clean_value(row.get('Category_name', 'N/A'))} "
                f"({clean_value(row.get('Category', 'N/A'))}); Policy: {clean_value(row.get('Policy_category_name', 'N/A'))}; "
                f"SSP: {clean_value(row.get('Ssp_family', 'N/A'))}; Tech: {clean_value(row.get('Technology_category_name', 'N/A'))}; "
                f"Peak CO2 year: {clean_value(row.get('Year of peak CO2 Emissions (Harm-Infilled)', 'N/A'))}; "
                f"Net-zero CO2 year: {clean_value(row.get('Year of netzero CO2 emissions (Harm-Infilled) Table SPM2', 'N/A'))}; "
                f"Median peak warming: {clean_value(row.get('Median peak warming (MAGICCv7.5.3)', 'N/A'))}°C; "
                f"Cumulative CO2 2020-2100: {clean_value(row.get('Cumulative net CO2 (2020-2100, Gt CO2) (Harm-Infilled)', 'N/A'))} Gt CO₂."
            )
            entries.append(entry)

    summary_text = (
        "AR6 情景参考（摘自官方数据库，用于约束哪些才算有效情景）:\n"
        + "\n".join(entries)
        if entries
        else "AR6 reference list empty. Treat only SSP/RCP/Net Zero style scenarios as valid."
    )
    if len(summary_text) > 8000:
        summary_text = summary_text[:8000] + "\n...(AR6 reference truncated)..."
    return summary_text, lookup


AR6_REFERENCE_TEXT, AR6_LOOKUP = build_ar6_reference_summary()

PAGES_BASIC_INFO = 3
PAGES_ROUTER = 2
PAGE_RANGE_METHODS = (1, 10)
PAGES_FULL_EXTRACT = None


EXPERT_INSTRUCTIONS = {
    "1. 气候观测与历史归因专家": """
    [新职责]:
    Focus on extracting the Model Performance Validation. Map the 'Historical Forcing Input' (e.g., natural + anthropogenic) directly to the 'Simulated Historical Trend'. Flag any reported discrepancies between the model run and observed data.
    提取重点：模型回测（Hindcast）设定、观测数据偏差（Bias）、归因结论的具体数值。
    """,
    "2. 全球气候模型与预测专家": """
    [新职责]:
    STRICT MAPPING PROTOCOL: You must create pairs of {Scenario_Input: SSPx-y, Time_Horizon: Year, Result_Output: Value ± Uncertainty}. Never mix results from high-emission scenarios with low-emission narratives. Identifying the specific CMIP generation (CMIP5 vs CMIP6) is mandatory.
    提取重点：ECS/TCR 敏感度参数、特定情景下的物理变量（GSAT, GMST）、多模型集合的中位数及不确定性范围。
    """,
    "3. 碳循环与生物地球化学专家": """
    [新职责]:
    Identify the Biogeochemical Feedback Settings. Specify if the model run included or excluded specific cycles (e.g., 'coupled carbon-cycle'). Extract the calculated TCRE (Transient Climate Response to Cumulative Emissions) value.
    提取重点：累积碳排放与温升比例、冻土/植被反馈开关、气溶胶强迫对应的降温效应。
    """,
    "4. 水循环与冰冻圈专家": """
    [新职责]:
    Map 'Global Warming Level' (Input) to 'Cryosphere/Hydrological Response' (Output). Look for threshold behaviors. Explicitly extract the contribution of thermal expansion vs. ice melt to Sea Level Rise projections.
    提取重点：温升阈值与不可逆临界点、区域水文变化（如 ENSO 对降水的影响）。
    """,

    "5. 生态系统与生物多样性专家": """
    [新职责]:
    Extract the link between 'Climatic Drivers' (e.g., max temperature) and 'Ecological Response' (e.g., mortality rate). Differentiate between results assuming natural adaptation vs. no adaptation.
    提取重点：气候生态位模型设定、物种栖息地丧失百分比、有无适应措施的差异。
    """,
    "6. 食品、水与健康安全专家": """
    [新职责]:
    Identify the specific Process-Based Model settings (e.g., CO2 fertilization effect ON/OFF). Map specific warming degrees (e.g., +2°C) to specific sectoral impacts (e.g., Wheat yield -15%).
    提取重点：作物模型参数（CO2施肥/灌溉）、健康风险函数（湿球温度导致的工时损失）。
    """,
    "7. 城市、住区与基础设施专家": """
    [新职责]:
    Link 'Hazard Probability' (e.g., 1-in-100 year flood) to 'Exposure & Vulnerability' (e.g., damage in USD). Extract the effectiveness of specific adaptation scenarios (e.g., 'Adaptation Scenario A' reduces damage by X%).
    提取重点：灾害暴露模型（海平面+风暴潮）、硬防护 vs 基于自然的解决方案减损效果对比。
    """,
    "8. 贫困、生计与可持续发展专家": """
    [新职责]:
    Focus on the intersection of Socio-economic Status and Climate Impact. Extract data showing how impacts vary across different income quintiles or social groups under specific scenarios.
    提取重点：SSP 社会经济参数（人口/基尼系数）、气候冲击导致的贫困发生率或移民人数。
    """,
    "9. 区域风险与决策评估专家": """
    [新职责]:
    Extract data for constructing 'Burning Embers' diagrams. Strictly map 'Global Warming Levels' (1.5°C, 2°C, 3°C) to specific 'Risk Transitions' (Moderate to High) for specific regions.
    提取重点：关键风险（Key Risks）矩阵、特定区域（如地中海）的综合风险级联。
    """,

    "10. 能源系统转型专家": """
    [新职责]:
    Identify Technological Constraints (e.g., 'No CCS scenario'). Map 'Policy Ambition' (Input) to 'Energy Mix Transformation' (Output). Extract the exact year of 'Peak Oil/Gas/Coal' under each scenario.
    提取重点：IAMs 模型约束（核能/CCS）、碳价格路径对应的化石燃料淘汰速度。
    """,
    "11. 城市、工业与交通减排专家": """
    [新职责]:
    Focus on Sectoral Activity Data. Link 'Mitigation Levers' (e.g., 50% EV penetration) to 'Emission Intensity Reductions'. Strictly separate Demand-side measures from Supply-side measures.
    提取重点：部门减排潜力、基准情景 vs 政策情景的能耗差异（钢铁/水泥等）。
    """,
    "12. 土地利用与AFOLU专家": """
    [新职责]:
    Analyze the trade-offs in Land Use. Map 'Carbon Removal Targets' (Input) to 'Land Area Requirements' (Output). Extract emission factors used for agriculture (e.g., methane per head of cattle).
    提取重点：土地竞争情景（BECCS需求）、不同管理下的土地碳汇通量。
    """,
    "13. 政策、金融与国际合作专家": """
    [新职责]:
    Map 'Policy Instruments' (e.g., Global Carbon Price) to 'Macroeconomic Indicators' (e.g., GDP loss, Consumption change). Verify if the scenario assumes global cooperation or fragmented policies.
    提取重点：宏观经济模型（GDP损失/边际减排成本 MAC）、不同温控目标下的资金流向预测。
    """
}


EXPERT_REPORT_KNOWLEDGE = {
    "1. 气候观测与历史归因专家": """
    参考报告：AR6 WG1 + AR6 SYR + AR5 WG1 + AR4 WG1
    重点知识：
    - WG1 Chapter 2：气候系统已发生的变化、观测事实、趋势与证据强度
    - WG1 Chapter 3：人为影响归因、自然强迫与人为强迫的区分
    - SYR：人类活动已 unequivocally 造成变暖的总论
    - 历史补充：AR5/AR4 的检测归因表述、历史强迫口径、SRES/RCP 时代的常见框架
    提取时优先识别：历史强迫组合、归因实验、观测-模拟对照、归因结论的量化结果。
    """,
    "2. 全球气候模型与预测专家": """
    参考报告：AR6 WG1 + AR6 SYR + AR5 WG1 + AR4 WG1
    重点知识：
    - WG1 Chapter 4：基于情景的未来气候变化预测
    - WG1 Chapter 11：极端事件在不同情景下的变化
    - SYR：1.5°C/2°C/更高温升下的关键差异
    - 历史补充：AR5 的 RCP 框架、AR4 的 SRES/稳定化情景表述
    提取时优先识别：SSP/RCP/SRES、CMIP5/CMIP6、时间范围、温度/降水/极端事件等投影结果。
    """,
    "3. 碳循环与生物地球化学专家": """
    参考报告：AR6 WG1 + AR6 SYR + AR5 WG1 + AR4 WG1
    重点知识：
    - WG1 Chapter 5：全球碳循环与其他生物地球化学循环
    - WG1 Chapter 6：短寿命气候强迫因子
    - WG1 Chapter 7：能量预算、气候反馈与气候敏感度
    - 历史补充：AR5/AR4 的累计排放、辐射强迫与碳循环反馈框架
    提取时优先识别：碳循环反馈、TCRE、CO2/CH4/N2O 情景、累计排放与温升关系。
    """,
    "4. 水循环与冰冻圈专家": """
    参考报告：AR6 WG1 + AR6 SYR + AR5 WG1 + AR4 WG1
    重点知识：
    - WG1 Chapter 8：水循环变化
    - WG1 Chapter 9：海洋、冰冻圈与海平面变化
    - WG1 Chapter 12：区域风险评估所需的气候信息
    - 历史补充：AR5/AR4 的海平面、冰川、积雪、径流与区域投影表述
    提取时优先识别：海平面上升、冰川/冰盖、积雪、径流、洪旱、水文阈值。
    """,
    "5. 生态系统与生物多样性专家": """
    参考报告：AR6 WG2 + AR6 SYR + AR5 WG2 + AR4 WG2
    重点知识：
    - WG2 中 terrestrial / freshwater / ocean & coastal ecosystems 相关章节
    - 生态系统响应、物种分布、生态系统服务与风险
    - 历史补充：AR5/AR4 的生态脆弱性、适应能力与 impacts/adaptation framing
    提取时优先识别：气候驱动、栖息地变化、灭绝/迁移风险、适应与无适应对比。
    """,
    "6. 食品、水与健康安全专家": """
    参考报告：AR6 WG2 + AR6 SYR + AR5 WG2 + AR4 WG2
    重点知识：
    - WG2 中 food、water、health 相关章节
    - 粮食安全、水安全、健康影响与脆弱性
    - 历史补充：AR5/AR4 的农业减产、需水变化、热健康与疾病风险口径
    提取时优先识别：作物产量、灌溉需水、热健康、疾病风险、饮水与水资源压力。
    """,
    "7. 城市、住区与基础设施专家": """
    参考报告：AR6 WG2 + AR6 SYR + AR5 WG2 + AR4 WG2
    重点知识：
    - WG2 城市、住区与关键基础设施章节
    - 暴露、脆弱性、适应措施与减损效果
    - 历史补充：AR5/AR4 的城市适应、基础设施暴露与灾损评估框架
    提取时优先识别：城市热风险、洪水/风暴潮、基础设施损失、适应方案效果。
    """,
    "8. 贫困、生计与可持续发展专家": """
    参考报告：AR6 WG2 + AR6 SYR + AR5 WG2 + AR4 WG2
    重点知识：
    - WG2 贫困、生计、可持续发展、脆弱性章节
    - 不同社会群体的差异化影响
    - 历史补充：AR5/AR4 的脆弱性、生计、发展路径与公平性表述
    提取时优先识别：收入分组、脆弱群体、贫困发生率、迁移、生计冲击。
    """,
    "9. 区域风险与决策评估专家": """
    参考报告：AR6 WG2 + AR6 SYR + AR5 WG2 + AR4 WG2
    重点知识：
    - WG2 key risks、decision-making、各区域章节
    - Burning embers、风险转折与区域复合风险
    - 历史补充：AR5/AR4 的 key vulnerabilities、reasons for concern、区域 impacts 框架
    提取时优先识别：全球温升水平与区域风险跃迁、决策场景、风险矩阵。
    """,
    "10. 能源系统转型专家": """
    参考报告：AR6 WG3 + AR6 SYR + AR5 WG3 + AR4 WG3
    重点知识：
    - WG3 能源系统、减缓路径章节
    - 技术约束、能源结构转型、净零路径
    - 历史补充：AR5/AR4 的稳定化路径、技术选项、基准情景与减缓情景框架
    提取时优先识别：能源情景、技术组合、碳价、化石燃料退出年份、电力结构变化。
    """,
    "11. 城市、工业与交通减排专家": """
    参考报告：AR6 WG3 + AR6 SYR + AR5 WG3 + AR4 WG3
    重点知识：
    - WG3 transport、industry、buildings、demand-side mitigation 相关章节
    - 部门减排杠杆与需求侧/供给侧区分
    - 历史补充：AR5/AR4 的部门减排潜力、技术扩散与需求侧措施表述
    提取时优先识别：EV 渗透率、工业能效、建筑节能、部门排放强度变化。
    """,
    "12. 土地利用与AFOLU专家": """
    参考报告：AR6 WG3 + AR6 SYR + AR5 WG3 + AR4 WG3
    重点知识：
    - WG3 AFOLU、土地竞争、CDR 相关章节
    - 土地碳汇、BECCS、农业与林业减缓
    - 历史补充：AR5/AR4 的 LULUCF/AFOLU 口径、林业与农业减缓路径
    提取时优先识别：土地需求、碳移除目标、农业排放因子、AFOLU 管理情景。
    """,
    "13. 政策、金融与国际合作专家": """
    参考报告：AR6 WG3 + AR6 SYR + AR5 WG3 + AR4 WG3
    重点知识：
    - WG3 policy instruments、finance、enabling conditions、international cooperation 章节
    - 政策工具到宏观经济/资金流/合作格局的映射
    - 历史补充：AR5/AR4 的全球合作、区域碎片化、边际减排成本和政策工具框架
    提取时优先识别：碳价、政策包、GDP/消费变化、投资需求、合作与碎片化情景。
    """,
}


def build_expert_report_knowledge(expert_name: str) -> str:
    summary = EXPERT_REPORT_KNOWLEDGE.get(expert_name, "").strip()
    if not summary:
        return "无专属报告知识摘要，按通用气候情景抽取规则执行。"
    return (
        f"{summary}\n"
        f"本地报告目录：{IPCC_REPORTS_DIR}\n"
        "本地知识库优先顺序：AR6 为主，AR5/AR4 为历史补充；旧论文出现 SRES、RCP、稳定化路径等历史框架时，可结合历史报告理解。\n"
        "你应把上述报告知识当作外部领域知识库，用来理解该专家在 IPCC 中的工作边界、变量口径、风险框架和常见情景类型；"
        "但提取结果必须以当前论文原文为准，不得用报告内容替代论文证据。"
    )

# ================= Prompt definitions =================

# Step 1: extract basic metadata
PROMPT_BASIC_INFO = """(deprecated - merged into PROMPT_COMBINED_SCREEN)"""

# Step 1.5: scenario pre-screen (fast filter for papers without scenarios)
PROMPT_SCENARIO_SCREEN = """
You are a climate/environmental science expert. Based on the abstract, keywords, AND methods section below, determine whether this paper contains **scenario-based modeling or simulations**.

A "scenario" is a coherent set of assumptions about driving factors (emissions, meteorology, socio-economic activity, policy interventions, etc.) that affect atmospheric/environmental conditions. Scenarios can be about the future, the past, or hypothetical conditions.

**Step 1: Check the Methods section for numerical/model simulation signals:**
- Does the paper use numerical models, climate models, atmospheric models, IAMs, or any simulation tools?
- Does it mention model runs, simulations, forcing experiments, sensitivity tests, or scenario-driven analysis?
- Key indicators: GCM/RCM/WRF/CMAQ/CAMx runs, IAM models (TIMES/MESSAGE/GCAM/REMIND), ecosystem/hydrological/crop models, agent-based models, system dynamics models, LCA under scenarios

**Step 2: If modeling/simulation is found, check for scenario setups of ANY type:**
- Baseline/BAU scenarios (current trends projected forward)
- Policy/control scenarios (specific emission reduction measures, carbon tax, renewable targets)
- Sensitivity/perturbation scenarios (source contribution analysis, zero-out, emission scaling)
- Comprehensive pathway scenarios (SSP, RCP, SRES A1/A2/B1/B2)
- Management scenarios (land-use change, grazing intensity, irrigation strategies)
- Custom scenarios (e.g., "+2°C warming", "20% emission reduction", specific technology adoption)

Papers that LIKELY CONTAIN scenarios:
- Run models under different assumptions (even just baseline vs. one alternative)
- Compare outcomes across scenarios or against a reference case
- Use SSP, RCP, SRES, or any policy/management pathway
- Perform sensitivity analysis by perturbing emissions or boundary conditions
- Project future states OR simulate hypothetical/counterfactual conditions

Papers that LIKELY DO NOT contain scenarios:
- Pure observational/monitoring studies (only analyze measured data, no model simulations at all)
- Methodological papers (develop algorithms without applying them to any scenario)
- Review/meta-analysis papers (summarize others' findings without own model runs)
- Pure statistical analysis of historical data without any simulation component
- Laboratory experiments or field measurements without modeling

Return JSON: {"has_scenario": true/false, "confidence": "high"/"medium"/"low", "reason": "Brief one-sentence explanation"}
"""

# Merged prompt: one call covers metadata extraction, scenario pre-screen, expert routing and location extraction
PROMPT_COMBINED_SCREEN = """
You are a climate/environmental science expert. Read the following paper text (abstract + methods) and complete ALL four tasks in a single response.

### Task 1: Extract basic metadata
Extract the paper's metadata from the first pages:
- Title: paper title
- Authors: first author or first three authors
- Journal: publishing journal (usually in header/footer)
- DOI: if provided, otherwise "N/A"
- Year: publication year

### Task 2: Scenario screening
Determine whether this paper runs its OWN model/simulation driven by explicit scenario assumptions.

A paper HAS scenarios when it satisfies BOTH conditions:
1. **The paper runs a numerical model or simulation** (GCM, RCM, crop model, hydrological model, IAM, ecosystem model, etc.) — not just references or cites other models' results
2. **The model is driven by explicit scenario assumptions** (SSP, RCP, SRES, emission pathways, policy interventions, land-use changes, management strategies, etc.)

Mark has_scenario=TRUE: paper runs its own model under scenario assumptions and reports simulation outputs.

Mark has_scenario=FALSE for ALL of the following — even if the paper mentions "scenario" or "RCP" in passing:
- Pure observational/monitoring studies analyzing measured or satellite data without running a simulation model
- Statistical/econometric analysis of historical data (regression, correlation, time series analysis) — even if climate variables are used as predictors
- Review/meta-analysis papers summarizing others' results without own model runs
- Behavioral experiments, surveys, or social science studies without physical/numerical modeling
- Papers that only USE scenario outputs from other studies as input data (e.g., "under RCP8.5 projections from CMIP6...") but do not run their own scenario-driven model
- Pure methodology/algorithm papers without scenario application

### Task 3: Expert routing
Classify the paper into exactly ONE of these 13 expert categories:
=== WG I ===
1. 气候观测与历史归因专家
2. 全球气候模型与预测专家
3. 碳循环与生物地球化学专家
4. 水循环与冰冻圈专家
=== WG II ===
5. 生态系统与生物多样性专家
6. 食品、水与健康安全专家
7. 城市、住区与基础设施专家
8. 贫困、生计与可持续发展专家
9. 区域风险与决策评估专家
=== WG III ===
10. 能源系统转型专家
11. 城市、工业与交通减排专家
12. 土地利用与AFOLU专家
13. 政策、金融与国际合作专家

### Task 4: Study location
Find the geographic scope described in the Study Area or Methods section.
- Ignore author affiliations/addresses.
- Look for explicit scope (e.g., "Global", "China", "Yangtze River Delta").
- Pure theoretical or global models → "Global".

### Return strict JSON:
{
  "title": "...",
  "authors": "...",
  "journal": "...",
  "doi": "...",
  "pub_year": "...",
  "has_scenario": true/false,
  "screen_confidence": "high/medium/low",
  "screen_reason": "Brief one-sentence explanation",
  "expert_category": "Exact name from list above",
  "study_location": "..."
}
"""

# Legacy prompt kept for reference elsewhere (no longer called directly)
PROMPT_ROUTER = """(deprecated - merged into PROMPT_COMBINED_SCREEN)"""
PROMPT_LOCATION = """(deprecated - merged into PROMPT_COMBINED_SCREEN)"""

# Step 4: core extraction (strict exclusivity rules, AR6 mapping instructions, model details, time-series variables)
PROMPT_SCENARIO_BASE = """
### Task: Extract Scenario → Simulation Output Pairs with Provenance
You are analyzing a climate/environmental modeling paper. A "scenario" is a coherent set of assumptions about driving factors (emissions, meteorology, socio-economic activity, policy interventions, etc.) used to drive model simulations. Capture ALL valid scenario types and their directly linked simulation outputs.

**Valid scenario types include:**
- Comprehensive pathway scenarios (SSP, RCP, SRES A1/A2/B1/B2)
- Baseline/BAU scenarios (current trends as reference)
- Policy/control scenarios (emission reductions, carbon tax, technology adoption)
- Sensitivity/perturbation scenarios (source apportionment, zero-out, emission scaling)
- Management scenarios (land-use, grazing, irrigation, conservation)
- Custom scenarios (+2°C warming, 20% emission reduction, etc.)

**Role Context**: You are the **{expert_name}**.
**Study Location**: {location}.

### EXPERT REPORT KNOWLEDGE BASE
{expert_report_knowledge}

### AR6 REFERENCE SNAPSHOT (authoritative examples of valid scenarios)
{ar6_reference}

### MANDATORY SSP/RCP → IPCC AR6 MAPPING TABLE
When you detect any of the following scenario names, you MUST fill the AR6 fields using this mapping. Do NOT leave them as "N/A" if a match exists:

| Scenario    | AR6 Category | Temperature Target | Net Zero Year | Carbon Budget     |
|-------------|-------------|-------------------|---------------|-------------------|
| SSP1-1.9    | C1          | 1.5°C             | 2050          | ~500 Gt CO2       |
| SSP1-2.6    | C3          | 1.8°C             | 2070          | ~900 Gt CO2       |
| SSP2-4.5    | C5          | 2.7°C             | N/A           | N/A               |
| SSP3-7.0    | C7          | 3.6°C             | N/A           | N/A               |
| SSP5-3.4-OS | C2          | 1.6°C (overshoot) | 2055          | ~700 Gt CO2       |
| SSP4-3.4    | C4          | 2.1°C             | 2080          | N/A               |
| SSP4-6.0    | C6          | 2.9°C             | N/A           | N/A               |
| SSP5-8.5    | C8          | >4°C (4.4°C)      | N/A           | N/A               |
| RCP2.6      | C3 (equiv)  | ~1.8°C            | ~2070         | ~900 Gt CO2       |
| RCP4.5      | C5 (equiv)  | ~2.7°C            | N/A           | N/A               |
| RCP6.0      | C6 (equiv)  | ~3.1°C            | N/A           | N/A               |
| RCP8.5      | C8 (equiv)  | >4°C (4.3°C)      | N/A           | N/A               |

AR6 Categories Reference:
- C1: Limiting warming to 1.5°C with no or limited overshoot
- C2: Return warming to 1.5°C after high overshoot
- C3: Limiting warming to 2°C (>67% probability)
- C4: Limiting warming to 2°C (>50% probability)
- C5: Limiting warming to 2.5°C
- C6: Limiting warming to 3°C
- C7: Limiting warming to 4°C
- C8: Exceeding 4°C (Current Policies)

### STANDARD IPCC SCENARIO KNOWLEDGE BASE (use this to enrich scenario_description)
When the paper only mentions a scenario name without explaining its assumptions, use this authoritative knowledge to write a complete scenario_description that explains WHAT THE SCENARIO ASSUMES (not what the paper does):

**SRES Scenarios (IPCC TAR/AR4, 2000):**
- **A1FI**: Rapid economic growth, global population peaks mid-century then declines, new technologies, fossil-fuel intensive. ~970 ppm CO2 by 2100, ~4°C warming.
- **A1B**: Same as A1 but balanced energy mix. ~720 ppm CO2 by 2100, ~2.8°C warming.
- **A1T**: Same as A1 but non-fossil energy emphasis. ~580 ppm CO2 by 2100, ~2.4°C warming.
- **A2**: High population growth (~15 billion by 2100), regionally fragmented economic development, slow and divided technology change, heavy fossil fuel reliance. ~850 ppm CO2 by 2100, radiative forcing ~8.5 W/m², ~3.4°C warming (range 2.0-5.4°C).
- **B1**: Convergent world, global population peaks mid-century, rapid shift to service/information economy, clean and resource-efficient technologies. ~550 ppm CO2 by 2100, ~1.8°C warming.
- **B2**: Intermediate population/economic growth, local solutions to sustainability, less rapid technology change. ~600 ppm CO2 by 2100, ~2.4°C warming.

**RCP Scenarios (IPCC AR5, 2014) — radiative forcing pathways:**
- **RCP2.6**: Strong mitigation, peak then decline. Radiative forcing ~2.6 W/m² by 2100, CO2 ~420 ppm, ~1.8°C warming.
- **RCP4.5**: Stabilization without overshoot. ~4.5 W/m², CO2 ~540 ppm, ~2.7°C warming.
- **RCP6.0**: Stabilization without overshoot. ~6.0 W/m², CO2 ~670 ppm, ~3.1°C warming.
- **RCP8.5**: High baseline emissions, no climate policy. ~8.5 W/m², CO2 ~940 ppm, >4°C warming.

**SSP Scenarios (IPCC AR6, 2021) — combine socioeconomic narrative + RCP forcing:**
- **SSP1-1.9**: Sustainability path with strongest mitigation. ~1.5°C warming, net-zero ~2050.
- **SSP1-2.6**: Sustainability path with strong mitigation. ~1.8°C warming, net-zero ~2070.
- **SSP2-4.5**: Middle of the road, moderate mitigation. ~2.7°C warming.
- **SSP3-7.0**: Regional rivalry, high challenges to mitigation/adaptation, fragmented world. ~3.6°C warming.
- **SSP4-6.0**: Inequality, mixed mitigation. ~2.9°C warming.
- **SSP5-8.5**: Fossil-fueled development, high economic growth, high energy demand. >4°C warming, ~CO2 doubling by 2050.

**Instructions for scenario_description:**
- If paper provides detailed scenario assumptions, USE THE PAPER'S DESCRIPTION (preferred — most accurate to study context).
- If paper only names the scenario (e.g., "we used A2"), write a description COMBINING the standard assumptions above with any context from the paper. Example: "A2 is a high-emission SRES scenario assuming high population growth (~15 billion by 2100), regionally fragmented economic development, slow technology change, and heavy fossil fuel reliance, leading to ~850 ppm CO2 by 2100 and ~3.4°C global warming. The paper applies it to model bird functional diversity changes."
- For custom/non-standard scenarios, describe the specific assumptions stated in the paper.
- ALWAYS focus on scenario assumptions (climate/emission/policy/socioeconomic conditions), NOT on what the paper does with the scenario.

### YOUR SPECIFIC INSTRUCTIONS:
{expert_specific_instructions}

### STRICT EXCLUSION RULES (严格排他条件)
⛔ **DO NOT** extract any of the following as scenarios:
- **Methodologies**: Statistical downscaling methods, bias correction techniques, model calibration procedures, ensemble weighting approaches
- **Assessment Frameworks**: Vulnerability assessment frameworks, risk matrices without specific scenario inputs, indicator systems
- **Statistical Models**: Regression models, machine learning classifiers, trend detection methods that do not project future states
- **General Narratives**: Statements like "climate change will affect..." without specific forcing/policy/pathway assumptions
- **Literature Reviews**: Summaries of what other papers found, meta-analyses without their own scenario runs

✅ A valid scenario **MUST** contain:
- A specific set of assumptions about driving factors — can be forward-looking (future projections), retrospective (historical counterfactuals), or hypothetical (sensitivity tests)
- Examples: "Under SSP5-8.5 forcing...", "If carbon tax reaches $100/tCO2...", "With 20% emission reduction from transport sector...", "Under A2 SRES pathway...", "Baseline scenario assuming current policies continue..."
- A quantitative or semi-quantitative simulation output directly driven by those assumptions
- A time horizon or analysis period

### STRICT RULES
1. A scenario is valid if it defines specific assumptions that drive a model simulation. This includes baseline/BAU, policy interventions, sensitivity tests, and comprehensive pathways (SSP/RCP/SRES). The scenario does NOT need to be forward-looking — historical counterfactual or perturbation scenarios are equally valid. Ignore generic narrative statements or pure methodology descriptions.
2. Each scenario must have exactly one logically corresponding output. Do **not** mix the output from scenario A with the description of scenario B.
3. Extract the exact original sentences for both the scenario description and the simulation output. Provide **three sentences** each: the sentence containing the required information plus the immediately preceding and following sentences (split by period / full stop). If a neighboring sentence does not exist, omit it but never fabricate content.
4. Preserve causality in the structured summary: explicitly state how the scenario assumptions drive the simulated result.
5. **AR6 Mapping is MANDATORY**: If the paper uses any recognized SSP or RCP pathway, you MUST fill ar6_category, ar6_temperature_target, ar6_net_zero_year, and ar6_carbon_budget using the mapping table above. Only use "N/A" when the scenario is a custom/novel pathway not in the table.
6. **Extract the original scenario name exactly** as it appears in the paper (e.g., "SSP1-1.9", "RCP8.5", "Net Zero 2050", "1.5°C-compatible"), then map it to the AR6 classification.
7. **Model Details**: Extract the specific model name (e.g., MESSAGEix-GLOBIOM, GCAM 5.4, REMIND-MAgPIE), spatial resolution (Global / Continental / National / Regional / Grid-level with resolution), and simulation time span (e.g., 2015-2100).
8. **Key Variable Time Series**: For each scenario, extract specific numerical values of key variables at different time points. Focus on: GHG emissions, fossil fuel share, carbon price, grid emission factor, and co-benefit pollutant concentrations (e.g., PM2.5). Format each as an independent [Year - Variable - Change Value - Brief Context] tuple. Do NOT merge multi-year data into a single description.
9. **Decode generic labels**: If the paper names a scenario as "Scenario 1/2/3", "Case A/B", "Treatment I/II", etc., keep that as scenario_name but use scenario_description to spell out the concrete setup from the paper. The generic label alone is never enough.
10. **Only keep grounded outputs**: output_value must be an explicit value attributable to that exact scenario in the paper text, figure, or table. If the paper only gives qualitative comparisons for that scenario, do not invent a number.

### JSON OUTPUT (STRICT)
Return JSON only with this schema:
{{
  "scenarios": [
    {{
      "scenario_name": "Original name exactly as in paper (e.g., SSP5-8.5, RCP2.6, Net Zero 2050)",
      "model_name": "Specific model name (e.g., MESSAGEix-GLOBIOM v4.0, GCAM 5.4, REMIND 2.1)",
      "model_spatial_resolution": "Global / National (country name) / Regional (region name) / Grid (resolution, e.g., 0.5°×0.5°)",
      "model_time_span": "Simulation period (e.g., 2015-2100, 2020-2060)",
      "time_horizon": "Key projection year or period (e.g., 2050, 2081-2100)",
      "scenario_description": "Describe the SCENARIO's own assumptions — what climate/emission/policy/management conditions does it assume? If the paper uses a generic label like 'Scenario 1', decode that label into the concrete setup from the paper. Do NOT describe the paper's research purpose or methodology here.",
      "scenario_context_text": "Sentence_before. Target_sentence. Sentence_after.",
      "ar6_category": "MUST fill if SSP/RCP detected: C1-C8 from mapping table. Only 'N/A' for truly novel scenarios.",
      "ar6_temperature_target": "MUST fill if SSP/RCP detected: temperature target from mapping table.",
      "ar6_net_zero_year": "MUST fill if SSP/RCP detected: net-zero year from mapping table.",
      "ar6_carbon_budget": "MUST fill if SSP/RCP detected: carbon budget from mapping table.",
      "output_variable": "Key simulated metric (GMST, GDP loss, crop yield, PM2.5, etc.)",
      "output_value": "Specific QUANTITATIVE value with units (e.g., '+2.3°C', '-15%', '488.9 million t', '3200 km² reduction'). Must be a number explicitly supported for this exact scenario in the paper — do NOT write qualitative descriptions like 'increased' or 'decreased'.",
      "output_description": "Plain-language description of the model result directly caused by the scenario, grounded in the paper's evidence for that exact scenario.",
      "output_context_text": "Sentence_before. Target_sentence. Sentence_after.",
      "key_variables_timeseries": [
        {{
          "year": "2030",
          "variable": "Grid emission factor / GHG emissions / fossil fuel share / carbon price / PM2.5 etc.",
          "value": "Specific number with unit or percentage change (e.g., 下降30%, 450 g CO2/kWh, $80/tCO2)",
          "context": "Brief context explaining the driver (e.g., 煤电逐步退出, 碳市场扩大覆盖面)"
        }}
      ],
      "structured_summary": "One paragraph produced by you tying the scenario assumptions to the simulation output, including AR6 classification rationale."
    }}
  ]
}}

Do not add fields beyond those specified. Skip any scenario without both description and output context.
If NO valid scenarios are found (only methodologies/frameworks/reviews), return: {{"scenarios": []}}
"""

# ================= Multi-Agent Debate Prompts (All English) =================

DEBATE_ROUND1_SYSTEM_PROMPT = """You are a climate scenario data quality expert participating in a three-way debate review. Independently review the following extraction result across five dimensions:

1. **Scenario Validity**: Does this represent a TRUE scenario — a coherent set of assumptions about driving factors (emissions, meteorology, socio-economic activity, policy interventions, etc.) used to drive model simulations? Valid scenario types include:
   - **Comprehensive pathway scenarios**: SSP, RCP, SRES (A1/A2/B1/B2) — all equally valid regardless of framework age
   - **Baseline/BAU scenarios**: current trends projected forward as a reference case
   - **Policy/control scenarios**: specific emission reductions, carbon tax, renewable targets, clean technology adoption
   - **Sensitivity/perturbation scenarios**: source contribution analysis, zero-out tests, emission scaling experiments
   - **Management scenarios**: grazing intensity, irrigation strategies, land-use change, conservation measures
   - **Custom scenarios**: "+2°C warming", "20% emission reduction", specific technology pathways
   A scenario does NOT need to be part of AR6/SSP/RCP to be valid. It only needs to define specific assumptions that drive a model simulation. It must NOT be purely a methodology description, assessment framework, or general narrative without concrete assumptions.
   Additionally check: Does scenario_description describe the SCENARIO'S OWN ASSUMPTIONS (climate/emission/policy conditions), or does it merely restate the paper's research purpose? If it describes the paper's purpose (e.g., "We predicted distribution shifts..."), flag this as an issue requiring fix — the description should state what the scenario assumes (e.g., "RCP8.5 assumes radiative forcing of 8.5 W/m² by 2100..."). If scenario_name is generic ("Scenario 1/2/3", "Case A/B", etc.), verify that scenario_description decodes the concrete setup from the paper; if not, flag it.
2. **Output Correspondence**: Does the simulation output CAUSALLY correspond to the scenario? Is the output directly driven by the scenario assumptions?
   Mark output as **VALID** when:
   - The output IS a result of a model/simulation driven by the scenario, even if the extracted value has minor issues (wrong units, absolute vs. relative, slight mismatch between output_variable name and output_value). These are extraction imperfections — flag them in fix_instruction but do NOT mark output as invalid.
   - The output covers one scenario from a multi-scenario comparison — this is valid as long as the value belongs to the stated scenario.
   Mark output as **INVALID** only when:
   - The output is NOT from any simulation (e.g., purely observational data, literature citation, or expert opinion)
   - The output is from a DIFFERENT scenario than the one stated (e.g., output says RCP8.5 result but scenario says RCP2.6)
   - The output has NO meaningful connection to the scenario (completely wrong field extracted)
   Additionally check: Is output_value a SPECIFIC QUANTITATIVE number with units? Qualitative descriptions like "increased", "decreased", or "changed" are NOT acceptable — flag as issue requiring fix with instruction to extract actual numbers from the paper. Also verify that output_context_text uniquely supports this exact scenario rather than a generic comparison sentence reused across multiple scenarios.
3. **Summary Fidelity**: Does the structured summary faithfully reflect both the scenario setup and the simulation result without adding unsupported claims? IMPORTANT: If the output is invalid (dimension 2), the summary that describes or relies on that incorrect output MUST also be marked invalid — a summary cannot be "faithful" if it faithfully reproduces wrong information.
4. **AR6 Mapping Correctness**: If the scenario belongs to SSP/RCP framework, are the AR6 fields correctly filled? If the scenario is NOT an SSP/RCP scenario (e.g., SRES, custom policy, management scenario), AR6 fields should be "N/A" — this is correct and expected, NOT a reason to mark the scenario as invalid.
5. **Timeseries Quality**: Are the key variable timeseries entries in proper [Year - Variable - Value - Context] format with reasonable values?

IMPORTANT: Scenario validity (dimension 1) and AR6 mapping (dimension 4) are INDEPENDENT. A scenario can be fully valid even if it has no AR6 mapping.

For each dimension marked as invalid, you MUST provide a specific, actionable correction instruction explaining WHAT is wrong and HOW to fix it.

Return strict JSON:
{
    "scenario_valid": true/false,
    "scenario_reasoning": "Detailed argument about scenario validity",
    "output_valid": true/false,
    "output_reasoning": "Detailed argument about output correspondence",
    "output_fix_instruction": "If invalid: specific instruction for re-extraction from original text, else empty string",
    "summary_valid": true/false,
    "summary_reasoning": "Detailed argument about summary fidelity",
    "summary_fix_instruction": "If invalid: specific instruction for correction, else empty string",
    "ar6_mapping_valid": true/false,
    "ar6_reasoning": "Detailed argument about AR6 mapping",
    "ar6_fix_instruction": "If invalid: the correct AR6 values that should be used, else empty string",
    "timeseries_valid": true/false,
    "timeseries_reasoning": "Detailed argument about timeseries quality",
    "timeseries_fix_instruction": "If invalid: specific format/content corrections needed, else empty string"
}
"""

DEBATE_ROUND2_SYSTEM_PROMPT = """You are in Round 2 of a three-way debate review. You have now seen the independent assessments from the other two experts.

Carefully read their arguments, then:
1. State which specific points you AGREE with and why
2. State which specific points you DISAGREE with, providing stronger evidence
3. If another expert found an issue you missed, acknowledge it
4. Compile a FINAL list of all issues that need fixing (consolidated from all three reviewers)

Return strict JSON:
{
    "agreements": "Points you agree with and why",
    "disagreements": "Points you disagree with and your counter-arguments",
    "updated_judgment": {
        "scenario_valid": true/false,
        "output_valid": true/false,
        "summary_valid": true/false,
        "ar6_mapping_valid": true/false,
        "timeseries_valid": true/false,
        "final_reasoning": "Your final comprehensive reasoning after debate"
    },
    "consolidated_issues": [
        {
            "field": "output/summary/ar6/timeseries",
            "problem": "What is wrong",
            "fix_instruction": "Specific instruction to fix it"
        }
    ]
}
"""

REEXTRACT_FIX_PROMPT = """You previously extracted scenario data from this paper, but the multi-agent review panel found specific issues. The scenario itself is VALID — do NOT discard it. Instead, fix ONLY the identified problems by re-reading the original paper text below.

### ISSUES TO FIX:
{issues_list}

### ORIGINAL EXTRACTION (for reference):
{original_extraction}

### INSTRUCTIONS:
- Re-read the paper text carefully to find the correct information
- Fix ONLY the fields mentioned in the issues list
- Keep all other fields unchanged
- If scenario_name is generic (e.g., Scenario 1/2/3, Case A/B), expand it into the paper's concrete assumptions inside scenario_description
- For output correspondence issues: find the EXACT output that is causally linked to this scenario
- Use only numbers explicitly supported by the paper text, table, or figure for this scenario; do not infer values from qualitative comparisons
- For summary issues: rewrite the summary to faithfully reflect scenario→output causality
- For timeseries issues: extract proper [year, variable, value, context] tuples from the paper
- For AR6 mapping: use the correct mapping (SSP1-1.9→C1/1.5°C, SSP1-2.6→C3/1.8°C, SSP2-4.5→C5/2.7°C, SSP3-7.0→C7/3.6°C, SSP5-8.5→C8/>4°C, RCP2.6→C3/~1.8°C, RCP4.5→C5/~2.7°C, RCP6.0→C6/~3.1°C, RCP8.5→C8/>4°C)

Return the COMPLETE corrected scenario as JSON (same schema as original extraction, single scenario object, not wrapped in "scenarios" array):
"""


def truncate_text(text: str, limit: int = 2000) -> str:
    if not text:
        return ""
    text = text.strip()
    if len(text) <= limit:
        return text
    return text[:limit] + "...(truncated)"


def init_client():
    if not API_KEY:
        print("❌ 错误：未找到 OPENAI_API_KEY")
        exit()
    return OpenAI(api_key=API_KEY, base_url=BASE_URL)


def init_gemini_client():
    if not GEMINI_API_KEY:
        print("⚠️ Gemini 验证已禁用（缺少 GEMINI_API_KEY 或 GOOGLE_API_KEY）")
        return None
    try:
        client = OpenAI(api_key=GEMINI_API_KEY, base_url=GEMINI_BASE_URL)
        return client
    except Exception as exc:
        print(f"⚠️ Gemini client init failed: {exc}")
        return None


def init_kimi_client():
    if not KIMI_API_KEY:
        print("⚠️ Kimi 验证已禁用（缺少 KIMI_API_KEY）")
        return None
    try:
        client = OpenAI(api_key=KIMI_API_KEY, base_url=KIMI_BASE_URL)
        return client
    except Exception as exc:
        print(f"⚠️ Kimi client init failed: {exc}")
        return None


def init_gpt4o_client():
    if not GPT4O_API_KEY:
        print("⚠️ GPT-4o validation disabled (no API key)")
        return None
    try:
        client = OpenAI(api_key=GPT4O_API_KEY, base_url=GPT4O_BASE_URL)
        return client
    except Exception as exc:
        print(f"⚠️ GPT-4o client init failed: {exc}")
        return None


_quota_exhausted = threading.Event()


def call_model(client: Optional[OpenAI], model: str, system_prompt: str, user_content: str, max_retries: int = 3) -> dict:
    """通用模型调用函数，返回解析后的 JSON，失败自动重试"""
    if client is None:
        return {"error": "Client unavailable"}
    if _quota_exhausted.is_set():
        return {"error": "已停止：API 额度耗尽"}
    for attempt in range(1, max_retries + 1):
        try:

            temp = 1 if "kimi" in model.lower() else 0
            res = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt.strip()},
                    {"role": "user", "content": user_content}
                ],
                temperature=temp
            )
            content = res.choices[0].message.content
            try:
                return json.loads(content)
            except json.JSONDecodeError:
                json_match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', content)
                if json_match:
                    return json.loads(json_match.group(1))
                if attempt < max_retries:
                    time.sleep(2 * attempt)
                    continue
                return {"error": f"无法解析响应: {content[:300]}"}
        except Exception as exc:
            exc_str = str(exc)
            if "insufficient_user_quota" in exc_str or "quota" in exc_str.lower() or "insufficient balance" in exc_str.lower() or ("402" in exc_str and "balance" in exc_str.lower()):
                _quota_exhausted.set()
                print(f"\n❌ API 额度耗尽，正在停止所有任务... ({exc_str[:100]})")
                return {"error": f"API 额度耗尽: {exc_str}"}
            if "API key not valid" in exc_str or "invalid_authentication" in exc_str.lower() or "Invalid Authentication" in exc_str:
                _quota_exhausted.set()
                print(f"\n❌ API key 无效，正在停止所有任务... ({exc_str[:100]})")
                return {"error": f"API key 无效: {exc_str}"}
            if attempt < max_retries:
                time.sleep(2 * attempt)
                continue
            return {"error": f"调用失败: {exc}"}


def build_debate_payload(scenario: dict) -> str:
    """构建辩论审查的完整 payload"""
    payload = {
        "scenario_name": scenario.get("scenario_name"),
        "scenario_description": truncate_text(scenario.get("scenario_description")),
        "scenario_context_text": truncate_text(scenario.get("scenario_context_text")),
        "model_name": scenario.get("model_name"),
        "model_spatial_resolution": scenario.get("model_spatial_resolution", "N/A"),
        "model_time_span": scenario.get("model_time_span", "N/A"),
        "time_horizon": scenario.get("time_horizon"),
        "ar6_category": scenario.get("ar6_category", "N/A"),
        "ar6_temperature_target": scenario.get("ar6_temperature_target", "N/A"),
        "ar6_net_zero_year": scenario.get("ar6_net_zero_year", "N/A"),
        "ar6_carbon_budget": scenario.get("ar6_carbon_budget", "N/A"),
        "output_variable": scenario.get("output_variable"),
        "output_value": scenario.get("output_value"),
        "output_description": truncate_text(scenario.get("output_description")),
        "output_context_text": truncate_text(scenario.get("output_context_text")),
        "key_variables_timeseries": scenario.get("key_variables_timeseries", []),
        "structured_summary": truncate_text(scenario.get("structured_summary"), limit=2500),
    }
    return json.dumps(payload, ensure_ascii=False)


def collect_issues_from_debate(openai_r2: dict, gemini_r2: dict, kimi_r2: dict) -> List[dict]:
    """Collect all consolidated issues from Round 2 debate results."""
    all_issues = []
    seen = set()
    for r2 in [openai_r2, gemini_r2, kimi_r2]:
        if "error" in r2:
            continue
        issues = r2.get("consolidated_issues", [])
        if isinstance(issues, list):
            for issue in issues:
                if isinstance(issue, dict):
                    key = (issue.get("field", ""), issue.get("problem", ""))
                    if key not in seen:
                        seen.add(key)
                        all_issues.append(issue)
    return all_issues


def check_scenario_valid_consensus(openai_r2: dict, gemini_r2: dict, kimi_r2: dict) -> bool:
    """Check if majority agrees the scenario itself is valid."""
    votes = []
    for r2 in [openai_r2, gemini_r2, kimi_r2]:
        if "error" in r2:
            continue
        uj = r2.get("updated_judgment", r2)
        v = uj.get("scenario_valid")
        if isinstance(v, bool):
            votes.append(v)
    if not votes:
        return False
    return sum(votes) > len(votes) / 2


def get_majority_flag(openai_r2: dict, gemini_r2: dict, kimi_r2: dict, field: str) -> Optional[bool]:
    """Return majority boolean for a debate field, or None if no usable votes exist."""
    votes = []
    for r2 in [openai_r2, gemini_r2, kimi_r2]:
        if "error" in r2:
            continue
        uj = r2.get("updated_judgment", r2)
        v = uj.get(field)
        if isinstance(v, bool):
            votes.append(v)
    if not votes:
        return None
    return sum(votes) > len(votes) / 2


def issue_is_blocking(issue: dict) -> bool:
    field = str((issue or {}).get("field", "")).strip().lower()
    return field in {"scenario", "output", "summary", "timeseries", "ar6"}


def has_blocking_issues(issues: List[dict]) -> bool:
    return any(issue_is_blocking(issue) for issue in issues)


def summarize_debate_outcome(openai_r2: dict, gemini_r2: dict, kimi_r2: dict) -> dict:
    issues = collect_issues_from_debate(openai_r2, gemini_r2, kimi_r2)
    scenario_valid = check_scenario_valid_consensus(openai_r2, gemini_r2, kimi_r2)
    output_valid = get_majority_flag(openai_r2, gemini_r2, kimi_r2, "output_valid")
    summary_valid = get_majority_flag(openai_r2, gemini_r2, kimi_r2, "summary_valid")
    ar6_mapping_valid = get_majority_flag(openai_r2, gemini_r2, kimi_r2, "ar6_mapping_valid")
    timeseries_valid = get_majority_flag(openai_r2, gemini_r2, kimi_r2, "timeseries_valid")

    extraction_complete = (
        scenario_valid
        and output_valid is True
        and summary_valid is True
        and ar6_mapping_valid is not False
        and timeseries_valid is not False
        and not has_blocking_issues(issues)
    )

    return {
        "issues": issues,
        "scenario_valid": scenario_valid,
        "output_valid": output_valid,
        "summary_valid": summary_valid,
        "ar6_mapping_valid": ar6_mapping_valid,
        "timeseries_valid": timeseries_valid,
        "extraction_complete": extraction_complete,
    }


def auto_fix_scenario(extract_client: OpenAI, scenario: dict, issues: List[dict], paper_text: str) -> dict:
    """Send identified issues back to extraction model with original paper text to fix."""
    issues_text = "\n".join([
        f"- [{issue.get('field', '?')}] {issue.get('problem', '?')}\n  Fix: {issue.get('fix_instruction', 'N/A')}"
        for issue in issues
    ])
    original_json = json.dumps({
        k: v for k, v in scenario.items()
        if k != "debate_validation"
    }, ensure_ascii=False, indent=2)

    prompt = REEXTRACT_FIX_PROMPT.format(
        issues_list=issues_text,
        original_extraction=original_json
    )

    for attempt in range(1, 4):
        try:
            res = extract_client.chat.completions.create(
                model=OPENAI_EXTRACT_MODEL,
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": paper_text[:80000]}
                ],
                response_format={"type": "json_object"},
                temperature=0.1
            )
            fixed = json.loads(res.choices[0].message.content)
            return fixed
        except Exception as e:
            if "quota" in str(e).lower():
                _quota_exhausted.set()
                print(f"\n❌ API 额度耗尽，正在停止所有任务...")
                return {}
            if attempt < 3:
                time.sleep(2 * attempt)
                continue
            print(f"  Auto-fix failed: {e}")
            return {}


def run_debate_and_fix(
    extract_client: OpenAI,
    openai_client: Optional[OpenAI],
    gemini_client: Optional[OpenAI],
    kimi_client: Optional[OpenAI],
    scenario: dict,
    paper_text: str,
    allow_auto_fix: bool = True,
) -> dict:
    """
    Multi-agent debate with auto-fix:
    Round 1: Three models independently review
    Round 2: Each model sees others' views, debates, and compiles issues
    Auto-Fix: If scenario is valid but other aspects have issues, re-extract from paper
    """
    payload_str = build_debate_payload(scenario)


    with ThreadPoolExecutor(max_workers=3) as pool:
        f_openai_r1 = pool.submit(call_model, openai_client, GPT4O_MODEL, DEBATE_ROUND1_SYSTEM_PROMPT, payload_str)
        f_gemini_r1 = pool.submit(call_model, gemini_client, GEMINI_MODEL, DEBATE_ROUND1_SYSTEM_PROMPT, payload_str)
        f_kimi_r1 = pool.submit(call_model, kimi_client, KIMI_MODEL, DEBATE_ROUND1_SYSTEM_PROMPT, payload_str)
    openai_r1 = f_openai_r1.result()
    gemini_r1 = f_gemini_r1.result()
    kimi_r1 = f_kimi_r1.result()


    r1_results = [openai_r1, gemini_r1, kimi_r1]
    r1_ok = [r for r in r1_results if "error" not in r]
    if len(r1_ok) < 2:
        print(f"    Debate: Round 1 失败过多 ({3 - len(r1_ok)}/3)，跳过辩论")
        return {
            "round1": {"openai": openai_r1, "gemini": gemini_r1, "kimi": kimi_r1},
            "round2": {}, "scenario_valid": False, "issues_found": [],
            "auto_fix_applied": False, "fix_details": None,
            "debate_skipped": True, "skip_reason": "Round 1 失败过多"
        }


    def build_round2_input(my_name, my_r1, other1_name, other1_r1, other2_name, other2_r1):
        return json.dumps({
            "original_extraction": json.loads(payload_str) if isinstance(payload_str, str) else payload_str,
            "your_round1_judgment": my_r1,
            f"{other1_name}_round1_judgment": other1_r1,
            f"{other2_name}_round1_judgment": other2_r1,
        }, ensure_ascii=False)

    openai_r2_input = build_round2_input("OpenAI", openai_r1, "Gemini", gemini_r1, "Kimi", kimi_r1)
    gemini_r2_input = build_round2_input("Gemini", gemini_r1, "OpenAI", openai_r1, "Kimi", kimi_r1)
    kimi_r2_input = build_round2_input("Kimi", kimi_r1, "OpenAI", openai_r1, "Gemini", gemini_r1)

    with ThreadPoolExecutor(max_workers=3) as pool:
        f_openai_r2 = pool.submit(call_model, openai_client, GPT4O_MODEL, DEBATE_ROUND2_SYSTEM_PROMPT, openai_r2_input)
        f_gemini_r2 = pool.submit(call_model, gemini_client, GEMINI_MODEL, DEBATE_ROUND2_SYSTEM_PROMPT, gemini_r2_input)
        f_kimi_r2 = pool.submit(call_model, kimi_client, KIMI_MODEL, DEBATE_ROUND2_SYSTEM_PROMPT, kimi_r2_input)
    openai_r2 = f_openai_r2.result()
    gemini_r2 = f_gemini_r2.result()
    kimi_r2 = f_kimi_r2.result()

    # ===== Analyze Debate Outcome =====
    outcome = summarize_debate_outcome(openai_r2, gemini_r2, kimi_r2)
    scenario_valid = outcome["scenario_valid"]
    issues = outcome["issues"]

    debate_result = {
        "round1": {"openai": openai_r1, "gemini": gemini_r1, "kimi": kimi_r1},
        "round2": {"openai": openai_r2, "gemini": gemini_r2, "kimi": kimi_r2},
        "scenario_valid": scenario_valid,
        "output_valid": outcome["output_valid"],
        "summary_valid": outcome["summary_valid"],
        "ar6_mapping_valid": outcome["ar6_mapping_valid"],
        "timeseries_valid": outcome["timeseries_valid"],
        "extraction_complete": outcome["extraction_complete"],
        "issues_found": issues,
        "auto_fix_applied": False,
        "fix_details": None,
    }

    # ===== Auto-Fix: If scenario valid but issues found, go back to paper =====
    if scenario_valid and issues and allow_auto_fix:
        print(f"    Debate: scenario valid but {len(issues)} issue(s) found. Auto-fixing...")
        fixed_data = auto_fix_scenario(extract_client, scenario, issues, paper_text)
        if fixed_data:
            # Merge fixed fields into scenario (only update fields that were actually fixed)
            for key, val in fixed_data.items():
                if key in scenario and val and str(val).strip() and str(val).strip() != "N/A":
                    scenario[key] = val
            debate_result["auto_fix_applied"] = True
            debate_result["fix_details"] = {
                "issues_sent": issues,
                "fields_updated": list(fixed_data.keys()),
            }
            # Re-run debate once on the fixed extraction. If critical issues remain, mark invalid.
            rerun_result = run_debate_and_fix(
                extract_client,
                openai_client,
                gemini_client,
                kimi_client,
                scenario,
                paper_text,
                allow_auto_fix=False,
            )
            rerun_result["auto_fix_applied"] = True
            rerun_result["fix_details"] = debate_result["fix_details"]
            return rerun_result
    elif not scenario_valid:
        print(f"    Debate: scenario INVALID — will be flagged, not fixed.")
    elif scenario_valid and not debate_result["extraction_complete"]:
        print(f"    Debate: scenario setup valid but some fields remain incomplete — keep result and record issues.")

    return debate_result


def apply_debate_and_fix(
    extract_client: OpenAI,
    openai_client: Optional[OpenAI],
    gemini_client: Optional[OpenAI],
    kimi_client: Optional[OpenAI],
    scenarios_data: dict,
    paper_text: str
) -> None:
    """Multi-agent debate + auto-fix for all scenarios in a paper."""
    if not isinstance(scenarios_data, dict):
        return
    scenarios = scenarios_data.get("scenarios")
    if not isinstance(scenarios, list):
        return

    for scenario in scenarios:
        if _quota_exhausted.is_set():
            return
        debate_result = run_debate_and_fix(
            extract_client, openai_client, gemini_client, kimi_client,
            scenario, paper_text
        )
        scenario["debate_validation"] = debate_result


def match_ar6_info(scenario_name: Optional[str]) -> Optional[Dict[str, Any]]:
    if not scenario_name:
        return None
    norm = normalize_key(scenario_name)
    if norm in AR6_LOOKUP:
        return AR6_LOOKUP[norm]
    if norm:
        for key, info in AR6_LOOKUP.items():
            if norm in key or key in norm:
                return info
    if norm in HARDCODED_AR6_MAPPING:
        return HARDCODED_AR6_MAPPING[norm]
    if norm:
        for key, info in HARDCODED_AR6_MAPPING.items():
            if norm in key or key in norm:
                return info
    return None


def enrich_scenarios_with_ar6(scenarios_data: dict) -> None:
    if not isinstance(scenarios_data, dict):
        return
    scenarios = scenarios_data.get("scenarios")
    if not isinstance(scenarios, list):
        return

    for scenario in scenarios:
        info = match_ar6_info(scenario.get("scenario_name"))
        if info:
            is_hardcoded = "ar6_category" in info
            for field, info_key in [
                ("ar6_category", "ar6_category"),
                ("ar6_temperature_target", "ar6_temperature_target"),
                ("ar6_net_zero_year", "ar6_net_zero_year"),
                ("ar6_carbon_budget", "ar6_carbon_budget"),
            ]:
                if info_key in info:
                    val = clean_value(info[info_key])
                else:
                    alt_keys = {
                        "ar6_category": ["category_name", "category"],
                        "ar6_temperature_target": ["median_peak_warming"],
                        "ar6_net_zero_year": ["net_zero_year"],
                        "ar6_carbon_budget": ["carbon_budget"],
                    }
                    val = "N/A"
                    for ak in alt_keys.get(info_key, []):
                        v = info.get(ak)
                        if v and clean_value(v) != "N/A":
                            val = clean_value(v)
                            break

                if is_hardcoded and val != "N/A":
                    scenario[field] = val
                else:
                    current = scenario.get(field, "N/A")
                    if is_blank_like(current):
                        scenario[field] = val
        else:
            scenario.setdefault("ar6_category", "N/A")
            scenario.setdefault("ar6_temperature_target", "N/A")
            scenario.setdefault("ar6_net_zero_year", "N/A")
            scenario.setdefault("ar6_carbon_budget", "N/A")

def extract_text(pdf_path, start_page, end_page):
    """通用文本读取函数"""
    content = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            start = max(0, start_page)
            if end_page is None:
                end = total_pages
            else:
                end = min(total_pages, end_page)

            for i in range(start, end):
                text = pdf.pages[i].extract_text()
                if text:
                    clean_text = text.replace(chr(0), '')
                    content += f"--- Page {i+1} ---\n{clean_text}\n"
        return content
    except Exception as e:
        print(f"⚠️ 读取PDF出错: {e}")
        return ""

def step1_extract_basic_info(client, text):
    """(deprecated) 保留兼容性"""
    return call_model(client, OPENAI_EXTRACT_MODEL, PROMPT_BASIC_INFO, text[:6000])


def step_combined_screen(client, text):
    """合并版：一次调用完成基础信息+情景预筛选+专家路由+地点提取"""
    result = call_model(client, OPENAI_EXTRACT_MODEL, PROMPT_COMBINED_SCREEN, text[:15000])
    if "error" in result:
        return {
            "basic_info": {"title": "N/A", "authors": "N/A", "journal": "N/A", "doi": "N/A", "pub_year": "N/A"},
            "has_scenario": True,
            "screen_confidence": "low",
            "screen_reason": "screening failed",
            "expert": "N/A",
            "location": "N/A"
        }
    return {
        "basic_info": {
            "title": result.get("title", "N/A"),
            "authors": result.get("authors", "N/A"),
            "journal": result.get("journal", "N/A"),
            "doi": result.get("doi", "N/A"),
            "pub_year": result.get("pub_year", "N/A")
        },
        "has_scenario": result.get("has_scenario", True),
        "screen_confidence": result.get("screen_confidence", "low"),
        "screen_reason": result.get("screen_reason", ""),
        "expert": result.get("expert_category", "N/A"),
        "location": result.get("study_location", "N/A")
    }

def step1_5_screen_scenario(client, text):
    """Step 1.5: 快速预筛选——判断论文是否可能包含情景建模/预测"""
    try:
        res = client.chat.completions.create(
            model=OPENAI_EXTRACT_MODEL,
            messages=[
                {"role": "system", "content": PROMPT_SCENARIO_SCREEN},
                {"role": "user", "content": text[:15000]}
            ],
            response_format={"type": "json_object"},
            temperature=0.0
        )
        result = json.loads(res.choices[0].message.content)
        has_scenario = result.get("has_scenario", True)
        confidence = result.get("confidence", "low")
        reason = result.get("reason", "")
        return has_scenario, confidence, reason
    except:
        return True, "low", "screening failed"

def step2_route_expert(client, text):
    try:
        res = client.chat.completions.create(
            model=OPENAI_EXTRACT_MODEL,
            messages=[
                {"role": "system", "content": PROMPT_ROUTER},
                {"role": "user", "content": text[:6000]}
            ],
            response_format={"type": "json_object"},
            temperature=0.0
        )
        expert = json.loads(res.choices[0].message.content).get("expert_category", "2. 全球气候模型与预测专家")
        if expert not in EXPERT_INSTRUCTIONS:
            return "2. 全球气候模型与预测专家"
        return expert
    except:
        return "2. 全球气候模型与预测专家"

def step3_find_location(client, text):
    try:
        res = client.chat.completions.create(
            model=OPENAI_EXTRACT_MODEL,
            messages=[
                {"role": "system", "content": PROMPT_LOCATION},
                {"role": "user", "content": f"请寻找 Study Area 或 Methods 章节:\n{text[:20000]}"}
            ],
            response_format={"type": "json_object"},
            temperature=0.0
        )
        return json.loads(res.choices[0].message.content).get("study_location", "Global")
    except:
        return "Global"

def step4_extract_scenarios(client, text, expert_name, location):
    specific_instruction = EXPERT_INSTRUCTIONS.get(expert_name, "")
    expert_report_knowledge = build_expert_report_knowledge(expert_name)


    final_prompt = PROMPT_SCENARIO_BASE.format(
        expert_name=expert_name,
        location=location,
        expert_specific_instructions=specific_instruction,
        expert_report_knowledge=expert_report_knowledge,
        ar6_reference=AR6_REFERENCE_TEXT
    )

    try:
        res = client.chat.completions.create(
            model=OPENAI_EXTRACT_MODEL,
            messages=[
                {"role": "system", "content": final_prompt},
                {
                    "role": "user",
                    "content": textwrap.dedent(
                        """
                        请阅读以下论文内容（尤其是情景设定与模拟结果），严格按照要求提取：
                        """
                    ).strip()
                },
                {"role": "user", "content": text[:80000]}
            ],
            response_format={"type": "json_object"},
            temperature=0.1
        )
        return json.loads(res.choices[0].message.content)
    except Exception as e:
        print(f"Scenario Extraction Error: {e}")
        return {}

def pdf_cache_key(pdf_path: str) -> str:
    """Generate a stable cache key from PDF path to handle Zotero's random subdirs.
    Uses the parent dir name (Zotero's 8-char hash) + filename for uniqueness."""
    parent = os.path.basename(os.path.dirname(pdf_path))
    fname = os.path.basename(pdf_path)
    # If parent is generic (like 'PDF' or 'storage'), fall back to hash of full path
    if len(parent) < 4 or parent in ("PDF", "storage", "pdf"):
        return hashlib.md5(pdf_path.encode()).hexdigest()[:12] + "_" + fname
    return parent + "_" + fname


def save_json_detail(item, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    cache_key = item.get("cache_key", item["file_name"].replace("/", "_"))
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', cache_key)
    out_path = os.path.join(output_dir, f"{safe_name}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(item, f, ensure_ascii=False, indent=2)


def format_debate_cell(debate_result: Optional[dict], model_name: str, round_key: str = "round1") -> str:
    """格式化辩论结果为单元格文本"""
    if not debate_result:
        return "Not Checked"
    round_data = debate_result.get(round_key, {})
    model_data = round_data.get(model_name, {})
    if "error" in model_data:
        return f"Error: {model_data['error']}"

    if round_key == "round1":
        overall_flag = model_data.get("scenario_valid")
        if not isinstance(overall_flag, bool):
            overall_flag = model_data.get("overall_valid")
        overall = "Valid" if overall_flag else "Invalid"
        reasoning = model_data.get("overall_reasoning") or model_data.get("final_reasoning", "")
        return f"{overall}: {reasoning}".strip()
    else:  # round2
        uj = model_data.get("updated_judgment", model_data)
        overall_flag = uj.get("scenario_valid")
        if not isinstance(overall_flag, bool):
            overall_flag = uj.get("overall_valid")
        overall = "Valid" if overall_flag else "Invalid"
        reasoning = uj.get("final_reasoning", "")
        agreements = model_data.get("agreements", "")
        disagreements = model_data.get("disagreements", "")
        parts = [f"{overall}"]
        if reasoning:
            parts.append(f"Reasoning: {reasoning}")
        if agreements:
            parts.append(f"Agrees: {agreements}")
        if disagreements:
            parts.append(f"Disagrees: {disagreements}")
        return " | ".join(parts)


def format_consensus_cell(debate_result: Optional[dict]) -> str:
    """Format debate final outcome."""
    if not debate_result:
        return "Not Checked"
    scenario_valid = debate_result.get("scenario_valid", False)
    issues = debate_result.get("issues_found", [])
    auto_fixed = debate_result.get("auto_fix_applied", False)

    if not scenario_valid:
        return "INVALID SCENARIO (not a true scenario)"
    if not issues:
        return "VALID (no issues found)"
    if auto_fixed:
        issue_fields = [str(i.get("field") or "?") for i in issues]
        return f"VALID + AUTO-FIXED ({', '.join(issue_fields)})"
    return f"VALID but {len(issues)} unfixed issue(s)"


def format_timeseries_cell(timeseries: Optional[list]) -> str:
    """格式化时间序列数据为单元格文本"""
    if not timeseries or not isinstance(timeseries, list):
        return "N/A"
    parts = []
    for item in timeseries:
        if isinstance(item, dict):
            year = item.get("year", "?")
            var = item.get("variable", "?")
            val = item.get("value", "?")
            ctx = item.get("context", "")
            parts.append(f"[{year}] {var}: {val} ({ctx})")
    return " || ".join(parts) if parts else "N/A"


def save_to_csv(results, filename):
    """Flatten structured scenario data into the required CSV columns."""
    headers = [
        "File Name",
        "Expert Category",
        "Title",
        "Authors",
        "Journal",
        "DOI",
        "Year",
        "Study Location",
        "Scenario Name",
        "Scenario Description",
        "Model Name",
        "Model Spatial Resolution",
        "Model Time Span",
        "AR6 Category",
        "AR6 Temperature Target / Peak Warming",
        "AR6 Net Zero Year",
        "AR6 Carbon Budget",
        "Scenario Context (Original Text)",
        "Output (Model Result)",
        "Output Context (Original Text)",
        "Key Variables Timeseries",
        "Structured Summary",
        "OpenAI Round1 Position",
        "Gemini Round1 Position",
        "Kimi Round1 Position",
        "OpenAI Round2 Position",
        "Gemini Round2 Position",
        "Kimi Round2 Position",
        "Debate Outcome",
        "Issues Found",
        "Auto-Fix Applied",
    ]

    rows = []
    for item in results:
        basic = item.get('basic_info', {})
        scenarios_block = item.get('scenarios_data') or {}
        scenarios = scenarios_block.get('scenarios', []) if isinstance(scenarios_block, dict) else []

        screen = item.get('scenario_screen', {})
        if screen.get('has_scenario') is False:
            scenario_label = f"No Scenario (filtered: {screen.get('reason', 'N/A')})"
        else:
            scenario_label = "No Valid Scenario"

        empty_row = [
            item.get('file_name', 'N/A'),
            item.get('expert', 'N/A'),
            basic.get('title', 'N/A'),
            basic.get('authors', 'N/A'),
            basic.get('journal', 'N/A'),
            basic.get('doi', 'N/A'),
            basic.get('pub_year', 'N/A'),
            item.get('location', 'N/A'),
            scenario_label,
            "N/A", "N/A", "N/A", "N/A",  # Description, Model details
            "N/A", "N/A", "N/A", "N/A",  # AR6 fields
            "N/A", "N/A", "N/A",  # Context + Output
            "N/A", "N/A",  # Timeseries + Summary
            "N/A", "N/A", "N/A",  # Round1
            "N/A", "N/A", "N/A",  # Round2
            "N/A",  # Debate Outcome
            "N/A",  # Issues Found
            "N/A",  # Auto-Fix
        ]

        if not scenarios:
            rows.append(empty_row)
            continue

        for s in scenarios:
            scenario_name = s.get('scenario_name', 'N/A')
            scenario_desc = s.get('scenario_description', 'N/A')
            model_name = s.get('model_name', 'N/A')
            model_spatial = s.get('model_spatial_resolution', 'N/A')
            model_time_span = s.get('model_time_span', 'N/A')
            time_horizon = s.get('time_horizon')

            desc_parts = [scenario_desc]
            if time_horizon:
                desc_parts.append(f"Time Horizon: {time_horizon}")
            scenario_desc_full = " | ".join([p for p in desc_parts if p and p != 'N/A']) or 'N/A'

            output_variable = s.get('output_variable', '')
            output_value = s.get('output_value', '')
            output_description = s.get('output_description', '')
            output_parts = [str(output_variable), str(output_value), str(output_description)]
            output_full = " | ".join([p for p in output_parts if p and p != 'N/A' and p != 'nan' and p]) or 'N/A'

            debate = s.get('debate_validation')

            rows.append([
                item.get('file_name', 'N/A'),
                item.get('expert', 'N/A'),
                basic.get('title', 'N/A'),
                basic.get('authors', 'N/A'),
                basic.get('journal', 'N/A'),
                basic.get('doi', 'N/A'),
                basic.get('pub_year', 'N/A'),
                item.get('location', 'N/A'),
                scenario_name,
                scenario_desc_full,
                model_name,
                model_spatial,
                model_time_span,
                s.get('ar6_category', 'N/A'),
                s.get('ar6_temperature_target', 'N/A'),
                s.get('ar6_net_zero_year', 'N/A'),
                s.get('ar6_carbon_budget', 'N/A'),
                s.get('scenario_context_text', 'N/A'),
                output_full,
                s.get('output_context_text', 'N/A'),
                format_timeseries_cell(s.get('key_variables_timeseries')),
                s.get('structured_summary', 'N/A'),
                format_debate_cell(debate, "openai", "round1"),
                format_debate_cell(debate, "gemini", "round1"),
                format_debate_cell(debate, "kimi", "round1"),
                format_debate_cell(debate, "openai", "round2"),
                format_debate_cell(debate, "gemini", "round2"),
                format_debate_cell(debate, "kimi", "round2"),
                format_consensus_cell(debate),
                # Issues found
                " | ".join([
                    f"[{i.get('field','?')}] {i.get('problem','?')}"
                    for i in (debate or {}).get("issues_found", [])
                ]) or "None",
                # Auto-fix applied
                "Yes" if (debate or {}).get("auto_fix_applied") else "No",
            ])


    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        red_font = Font(color="FF0000")
        scenario_col_idx = headers.index("Scenario Name") + 1
        for row_data in rows:
            ws.append(row_data)
            cell = ws.cell(row=ws.max_row, column=scenario_col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("No Scenario (filtered"):
                cell.font = red_font
        xlsx_path = filename.rsplit(".", 1)[0] + ".xlsx"
        wb.save(xlsx_path)
    except Exception as e:
        print(f"⚠️ 保存 xlsx 失败: {e}")


def main():
    client = init_client()
    gemini_client = init_gemini_client()
    kimi_client = init_kimi_client()
    gpt4o_client = init_gpt4o_client()

    print("\n" + "=" * 50)
    print("📋 模型配置:")
    print(f"   提取: OpenAI ({OPENAI_EXTRACT_MODEL})")
    if gpt4o_client:
        print(f"   辩论1: OpenAI ({GPT4O_MODEL}) ✅")
    else:
        print("   辩论1: OpenAI ❌ (已禁用)")
    if gemini_client:
        print(f"   辩论2: Gemini ({GEMINI_MODEL}) ✅")
    else:
        print("   辩论2: Gemini ❌ (已禁用)")
    if kimi_client:
        print(f"   辩论3: Kimi ({KIMI_MODEL}) ✅")
    else:
        print("   辩论3: Kimi ❌ (已禁用)")
    print("=" * 50 + "\n")

    print(f"📂 正在搜索目录 (递归): {FOLDER_PATH}")
    files = glob.glob(os.path.join(FOLDER_PATH, "**/*.pdf"), recursive=True)
    files.sort()
    if MAX_FILES_OVERRIDE > 0:
        print(f"⚙️ 仅提取前 {MAX_FILES_OVERRIDE} 篇 PDF 进行测试 (SCENARIO_MAX_FILES)")
        files = files[:MAX_FILES_OVERRIDE]

    if not files:
        print(f"❌ 未找到文件，请检查路径: {FOLDER_PATH}")
        return

    all_results = []
    skipped = 0
    filtered = 0


    os.makedirs(OUTPUT_DIR, exist_ok=True)
    existing_keys = set()
    need_debate_items = []
    for jf in glob.glob(os.path.join(OUTPUT_DIR, "*.json")):
        try:
            with open(jf, "r", encoding="utf-8") as f:
                cached = json.load(f)
            existing_keys.add(cached.get("cache_key", cached.get("file_name", "")))
            sd = cached.get("scenarios_data", {})
            scenarios = sd.get("scenarios", []) if isinstance(sd, dict) else []
            debate_missing = False
            if scenarios:
                for s in scenarios:
                    dv = s.get("debate_validation")
                    if not dv or dv.get("debate_skipped"):
                        debate_missing = True
                        break
                    for rd in ["round1", "round2"]:
                        for mdl in ["openai", "gemini", "kimi"]:
                            if "error" in dv.get(rd, {}).get(mdl, {}):
                                debate_missing = True
                                break
                        if debate_missing:
                            break
                    if debate_missing:
                        break
            if debate_missing:
                need_debate_items.append((jf, cached))
            else:
                all_results.append(cached)
        except Exception:
            pass


    file_keys = {pdf_path: pdf_cache_key(pdf_path) for pdf_path in files}
    to_process_list = [p for p in files if file_keys[p] not in existing_keys]
    print(f"🚀 找到 {len(files)} 个 PDF 文件，已有 {len(existing_keys)} 篇缓存结果")
    print(f"   需要补跑辩论: {len(need_debate_items)} 篇")
    print(f"   待全新处理: {len(to_process_list)} 篇，并发数: {MAX_WORKERS}")

    if need_debate_items:
        print(f"\n🔄 开始补跑辩论 ({len(need_debate_items)} 篇)...")
        debate_fixed = 0
        debate_failed = 0

        fname_to_path = {}
        for p in files:
            fname_to_path[os.path.basename(p)] = p

        def debate_one(item_tuple):
            jf_path, cached = item_tuple
            if _quota_exhausted.is_set():
                return None
            try:
                sd = cached.get("scenarios_data", {})
                fname = cached.get("file_name", "")
                pdf_path = fname_to_path.get(fname)
                text_full = extract_text(pdf_path, 0, PAGES_FULL_EXTRACT) if pdf_path else ""

                apply_debate_and_fix(client, gpt4o_client, gemini_client, kimi_client, sd, text_full)

                scenarios = sd.get("scenarios", [])
                all_ok = all(
                    s.get("debate_validation") and not s["debate_validation"].get("debate_skipped")
                    for s in scenarios
                )
                if all_ok and not _quota_exhausted.is_set():
                    cached["scenarios_data"] = sd
                    with open(jf_path, "w", encoding="utf-8") as f:
                        json.dump(cached, f, ensure_ascii=False, indent=2)
                    return cached
                return None
            except Exception as e:
                print(f"  ⚠️ 补辩论失败 {cached.get('file_name', '?')}: {e}")
                return None

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(debate_one, item): item for item in need_debate_items}
            for future in tqdm(as_completed(futures), total=len(need_debate_items), desc="补跑辩论"):
                result = future.result()
                if result:
                    all_results.append(result)
                    debate_fixed += 1
                else:
                    debate_failed += 1
        print(f"   补辩论完成: 成功 {debate_fixed}, 失败 {debate_failed}")

    lock = threading.Lock()
    counters = {"filtered": 0, "processed": 0, "errors": 0}

    def process_one_paper(pdf_path):
        """处理单篇论文（线程安全）"""
        if _quota_exhausted.is_set():
            return None
        fname = os.path.basename(pdf_path)
        cache_key = file_keys[pdf_path]

        try:

            text_screen = extract_text(pdf_path, 0, PAGE_RANGE_METHODS[1])
            if not text_screen:
                return None

            combined = step_combined_screen(client, text_screen)
            basic_info = combined["basic_info"]
            has_scenario = combined["has_scenario"]
            screen_confidence = combined["screen_confidence"]
            screen_reason = combined["screen_reason"]
            expert = combined["expert"]
            location = combined["location"]

            if not has_scenario and screen_confidence in ("high", "medium"):
                item = {
                    "file_name": fname,
                    "cache_key": cache_key,
                    "basic_info": basic_info,
                    "expert": "N/A (filtered)",
                    "location": location,
                    "scenarios_data": {"scenarios": []},
                    "scenario_screen": {
                        "has_scenario": False,
                        "confidence": screen_confidence,
                        "reason": screen_reason
                    }
                }
                save_json_detail(item, OUTPUT_DIR)
                with lock:
                    counters["filtered"] += 1
                return item

            # --- Stage 4: Scenario Extraction (OpenAI) ---
            text_full = extract_text(pdf_path, 0, PAGES_FULL_EXTRACT)
            scenarios_data = step4_extract_scenarios(client, text_full, expert, location)
            enrich_scenarios_with_ar6(scenarios_data)

            # --- Stage 5: Multi-Agent Debate + Auto-Fix (OpenAI + Gemini + Kimi) ---
            apply_debate_and_fix(client, gpt4o_client, gemini_client, kimi_client, scenarios_data, text_full)

            if _quota_exhausted.is_set():
                return None

            item = {
                "file_name": fname,
                "cache_key": cache_key,
                "basic_info": basic_info,
                "expert": expert,
                "location": location,
                "scenarios_data": scenarios_data,
                "scenario_screen": {
                    "has_scenario": True,
                    "confidence": screen_confidence,
                    "reason": screen_reason
                }
            }
            save_json_detail(item, OUTPUT_DIR)
            with lock:
                counters["processed"] += 1
            return item

        except Exception as e:
            print(f"⚠️ 处理失败 {fname}: {e}")
            with lock:
                counters["errors"] += 1
            return None

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_one_paper, p): p for p in to_process_list}
        pbar = tqdm(total=len(to_process_list), desc="处理进度")

        for future in as_completed(futures):
            result = future.result()
            if result is not None:
                with lock:
                    all_results.append(result)
                    if SAVE_EVERY_N and (len(all_results) % SAVE_EVERY_N == 0):
                        save_to_csv(all_results, OUTPUT_CSV)
            pbar.update(1)

        pbar.close()

    save_to_csv(all_results, OUTPUT_CSV)
    skipped = len(existing_keys)
    print(f"\n✅ 处理完成！")
    print(f"   新提取: {counters['processed']} 篇，预筛选过滤: {counters['filtered']} 篇（不含情景），跳过已缓存: {skipped} 篇，错误: {counters['errors']} 篇")
    print(f"   共 {len(all_results)} 篇结果")
    print(f"   CSV: {os.path.abspath(OUTPUT_CSV)}")
    print(f"   JSON: {os.path.abspath(OUTPUT_DIR)}")
    print(f"   如需重新处理某篇，删除对应的 JSON 文件后重跑即可")

if __name__ == "__main__":
    main()
